from tkinter import *
from openpyxl import workbook, load_workbook



def play():
    tk = Tk()
    tk.title('CADASTRAR CERTIFICADO')
    tk.geometry('400x300')
    





    nome = Label(tk, text="Nome da Empresa:", font=("Arial", 10,'bold'), fg="black")
    nome.pack(side = 'top')

    campo = Entry(tk,bg='light blue',width=50, font=('Arial',10,'bold'))
    campo.pack(side = 'top')

    # configurações de Entrada
    def on_entry_change(event):
        conteudo = campo.get()
        campo.delete(0, END)
        campo.insert(0, conteudo.upper())

    email = Label(tk, text="Email para contato:", font=("Arial", 10,'bold'), fg="black")
    email.pack(side = 'top')

    campo_email = Entry(tk, bg= 'light blue', width=50, font=('Arial',10,'bold'))
    campo_email.pack(side='top')



    campo.pack()
    campo.bind("<KeyRelease>", on_entry_change)

    data = Label(tk, text="Vencimento do Certificado", font=("Arial", 10,'bold'), fg="black")
    data.pack(side='top')

    campo_data = Entry(tk, bg= 'light blue',width=50,font=('Arial',10,'bold'))
    campo_data.pack(pady=10)


    def preencher():
        wb = load_workbook(r'C:\Users\Audi3\Documents\Automacoes\Vencimento\Dataframe\BD VENCIMENTOS.xlsx')
        ws = wb['Planilha1']
        coluna_nome = 'A'
        coluna_email = 'B'
        coluna_data = 'C'
        nome_vlr = str( campo.get())
        email_vlr = str (campo_email.get())
        data_vlr = str(campo_data.get())
        nome = nome_vlr
        email = email_vlr
        data = data_vlr
        # Procure a primeira célula vazia na coluna
        for row in ws[coluna_nome]:
            if not row.value:
            # Preencha a célula vazia
                row.value =nome
                print(nome_vlr)
                break
        for row in ws[coluna_email]:
            if not row.value:
            # Preencha a célula vazia
                row.value =email
                print(email_vlr)
                break
        for row in ws[coluna_data]:
            if not row.value:
            # Preencha a célula vazia
                row.value =data
                print(data_vlr)
                break
        wb.save(r'C:\Users\Audi3\Documents\Automacoes\Vencimento\Dataframe\BD VENCIMENTOS.xlsx')
        print('salvo')
    botao = Button(text='Enviar', activebackground='black' , fg = 'black', activeforeground= 'black',font=('Arial',10,'bold'),command=preencher)
    botao.pack(side='top')


    tk.mainloop()

if __name__ == '__main__':
    play()
